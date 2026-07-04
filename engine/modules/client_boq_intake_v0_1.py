"""Phase 2 client BOQ workbook intake and validation foundation.

This module is intentionally standalone. It reads artificial .xlsx fixtures,
records source traceability, validates the minimum client BOQ fields, and does
not call pricing, export, CE backend, router, architect, validator, merge, or
runtime pipeline modules.
"""

from __future__ import annotations

import os
from decimal import Decimal, InvalidOperation
from typing import Any, Dict, List, Optional, Tuple

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter


REQUIRED_FIELDS = ("client_item_ref", "description", "unit", "quantity")

HEADER_ALIASES = {
    "client_item_ref": {
        "item",
        "item no",
        "item no.",
        "item ref",
        "item reference",
        "client item ref",
        "client item reference",
        "boq item",
        "boq ref",
    },
    "description": {
        "description",
        "item description",
        "scope description",
        "work description",
        "boq description",
    },
    "unit": {
        "unit",
        "uom",
        "unit of measure",
        "units",
    },
    "quantity": {
        "quantity",
        "qty",
        "quantities",
        "client quantity",
    },
}

EXCLUSION_HEADERS = {
    "exclude",
    "excluded",
    "row state",
    "row status",
    "include",
}

EXCLUSION_MARKERS = {"exclude", "excluded", "no", "n", "skip", "omit"}


def import_client_boq_workbook(path: str, worksheet_name: Optional[str] = None) -> Dict[str, Any]:
    """Read a client BOQ fixture workbook and return validation evidence."""
    source_file_ref = os.path.basename(os.fspath(path))
    result = _empty_result(source_file_ref)

    try:
        workbook = load_workbook(path, data_only=False, read_only=False)
    except Exception as exc:
        result["workbook_readable"] = False
        result["workbook_status"] = "fail"
        result["workbook_messages"].append(f"Workbook is not readable: {type(exc).__name__}.")
        return result

    result["workbook_readable"] = True
    result["workbook_status"] = "pass"
    result["sheets"] = [
        {"worksheet_name": sheet.title, "worksheet_index": index}
        for index, sheet in enumerate(workbook.worksheets, start=1)
    ]

    worksheet, selection_message = _select_worksheet(workbook, worksheet_name)
    if worksheet is None:
        result["workbook_status"] = "fail"
        result["workbook_messages"].append(selection_message)
        return result

    result["selected_worksheet"] = {
        "worksheet_name": worksheet.title,
        "worksheet_index": workbook.worksheets.index(worksheet) + 1,
    }

    structural_messages = _worksheet_structure_messages(worksheet)
    result["workbook_messages"].extend(structural_messages)

    header_row_number, headers = _find_header_row(worksheet)
    result["header_row_number"] = header_row_number
    if header_row_number is None:
        result["column_mapping_status"] = "fail"
        result["workbook_status"] = "fail"
        result["workbook_messages"].append("No header row found.")
        return result

    mapping, mapping_messages, mapping_status, exclusion_column = _map_columns(headers)
    result["column_mapping"] = mapping
    result["column_mapping_status"] = mapping_status
    result["column_mapping_messages"] = mapping_messages

    if mapping_status != "pass":
        result["workbook_status"] = "fail" if mapping_status == "fail" else "warning"
        result["workbook_messages"].append("Column mapping requires review before pricing readiness.")

    data_start = header_row_number + 1
    for row_number in range(data_start, worksheet.max_row + 1):
        row = _build_row_result(
            worksheet=worksheet,
            row_number=row_number,
            mapping=mapping,
            exclusion_column=exclusion_column,
            source_file_ref=source_file_ref,
        )
        if row is not None:
            result["rows"].append(row)

    if not result["rows"]:
        result["workbook_messages"].append("No BOQ data rows found.")
        result["workbook_status"] = "fail"

    if structural_messages and result["workbook_status"] == "pass":
        result["workbook_status"] = "warning"

    if any(row["validation_status"] == "fail" for row in result["rows"]):
        result["workbook_status"] = "fail"
    elif any(row["validation_status"] in {"warning", "excluded"} for row in result["rows"]):
        if result["workbook_status"] == "pass":
            result["workbook_status"] = "warning"

    return result


def _empty_result(source_file_ref: str) -> Dict[str, Any]:
    return {
        "source_file_ref": source_file_ref,
        "workbook_readable": False,
        "workbook_status": "unknown",
        "workbook_messages": [],
        "sheets": [],
        "selected_worksheet": None,
        "header_row_number": None,
        "column_mapping_status": "unknown",
        "column_mapping": {},
        "column_mapping_messages": [],
        "rows": [],
        "pricing_approval": False,
        "export_ready": False,
    }


def _select_worksheet(workbook: Any, worksheet_name: Optional[str]) -> Tuple[Any, str]:
    if worksheet_name:
        if worksheet_name in workbook.sheetnames:
            return workbook[worksheet_name], "Worksheet selected by explicit name."
        return None, f"Worksheet not found: {worksheet_name}."

    visible_sheets = [sheet for sheet in workbook.worksheets if sheet.sheet_state == "visible"]
    if not visible_sheets:
        return None, "No visible worksheet found."
    return visible_sheets[0], "First visible worksheet selected."


def _worksheet_structure_messages(worksheet: Any) -> List[str]:
    messages: List[str] = []

    hidden_rows = [
        index
        for index, dimension in worksheet.row_dimensions.items()
        if dimension.hidden
    ]
    hidden_columns = [
        column
        for column, dimension in worksheet.column_dimensions.items()
        if dimension.hidden
    ]

    if hidden_rows:
        messages.append(f"Hidden rows visible for review: {hidden_rows}.")
    if hidden_columns:
        messages.append(f"Hidden columns visible for review: {hidden_columns}.")
    if worksheet.merged_cells.ranges:
        ranges = [str(cell_range) for cell_range in worksheet.merged_cells.ranges]
        messages.append(f"Merged cells visible for review: {ranges}.")

    return messages


def _find_header_row(worksheet: Any) -> Tuple[Optional[int], Dict[int, str]]:
    max_scan = min(worksheet.max_row, 10)
    for row_number in range(1, max_scan + 1):
        headers: Dict[int, str] = {}
        for cell in worksheet[row_number]:
            value = _clean_text(cell.value)
            if value:
                headers[cell.column] = value
        if headers:
            return row_number, headers
    return None, {}


def _map_columns(
    headers: Dict[int, str]
) -> Tuple[Dict[str, Dict[str, Any]], List[str], str, Optional[int]]:
    normalized = {column: _normalize_header(header) for column, header in headers.items()}
    mapping: Dict[str, Dict[str, Any]] = {}
    messages: List[str] = []
    status = "pass"

    for field in REQUIRED_FIELDS:
        matches = [
            column
            for column, header in normalized.items()
            if header in HEADER_ALIASES[field]
        ]
        if not matches:
            status = "fail"
            messages.append(f"Missing required column for {field}.")
            continue
        if len(matches) > 1:
            if status != "fail":
                status = "review_required"
            labels = [headers[column] for column in matches]
            messages.append(f"Ambiguous column mapping for {field}: {labels}.")
        selected = matches[0]
        mapping[field] = {
            "column_index": selected,
            "column_letter": get_column_letter(selected),
            "header": headers[selected],
        }

    exclusion_matches = [
        column
        for column, header in normalized.items()
        if header in EXCLUSION_HEADERS
    ]
    exclusion_column = exclusion_matches[0] if exclusion_matches else None

    if not messages:
        messages.append("Required fixture columns mapped deterministically.")
    return mapping, messages, status, exclusion_column


def _build_row_result(
    worksheet: Any,
    row_number: int,
    mapping: Dict[str, Dict[str, Any]],
    exclusion_column: Optional[int],
    source_file_ref: str,
) -> Optional[Dict[str, Any]]:
    row_values = {
        field: _cell_value(worksheet, row_number, spec.get("column_index"))
        for field, spec in mapping.items()
    }
    if not any(_clean_text(value) for value in row_values.values()):
        return None

    messages: List[str] = []
    status = "pass"
    estimator_review_required = False
    exclusion_reason = None

    hidden_or_merged_messages = _row_structure_messages(worksheet, row_number, mapping)
    if hidden_or_merged_messages:
        status = "warning"
        estimator_review_required = True
        messages.extend(hidden_or_merged_messages)

    exclusion_value = _cell_value(worksheet, row_number, exclusion_column)
    if _is_excluded(exclusion_value):
        status = "excluded"
        exclusion_reason = "Fixture row marked as excluded."
        messages.append(exclusion_reason)

    description = row_values.get("description")
    unit = row_values.get("unit")
    quantity = row_values.get("quantity")

    if not _clean_text(description):
        status = "fail"
        estimator_review_required = True
        messages.append("Description is required and is blank or missing.")
    if not _clean_text(unit):
        status = "fail"
        estimator_review_required = True
        messages.append("Unit is required and is blank or missing.")

    quantity_status, quantity_message = _validate_quantity(quantity)
    if quantity_status != "pass":
        status = "fail"
        estimator_review_required = True
        messages.append(quantity_message)

    if not messages:
        messages.append("Required row fields passed import validation.")

    return {
        "source_file_ref": source_file_ref,
        "source_sheet_name": worksheet.title,
        "source_row_number": row_number,
        "client_item_ref": row_values.get("client_item_ref"),
        "description_original": description,
        "unit_original": unit,
        "quantity_original": quantity,
        "quantity_normalized": _normalize_quantity(quantity) if quantity_status == "pass" else None,
        "validation_status": status,
        "validation_messages": messages,
        "readiness_status": _readiness_status(status),
        "pricing_ready": False,
        "export_ready": False,
        "estimator_review_required": estimator_review_required,
        "exclusion_reason": exclusion_reason,
    }


def _row_structure_messages(
    worksheet: Any,
    row_number: int,
    mapping: Dict[str, Dict[str, Any]],
) -> List[str]:
    messages: List[str] = []
    row_dimension = worksheet.row_dimensions[row_number]
    if row_dimension.hidden:
        messages.append(f"Source row {row_number} is hidden and requires review.")

    for field, spec in mapping.items():
        column_index = spec["column_index"]
        column_letter = get_column_letter(column_index)
        if worksheet.column_dimensions[column_letter].hidden:
            messages.append(f"Mapped {field} column {column_letter} is hidden and requires review.")

        cell_ref = f"{column_letter}{row_number}"
        if _cell_is_merged(worksheet, cell_ref):
            messages.append(f"Mapped {field} cell {cell_ref} is merged and requires review.")

    return messages


def _cell_is_merged(worksheet: Any, cell_ref: str) -> bool:
    for merged_range in worksheet.merged_cells.ranges:
        if cell_ref in merged_range:
            return True
    return False


def _cell_value(worksheet: Any, row_number: int, column_index: Optional[int]) -> Any:
    if column_index is None:
        return None
    return worksheet.cell(row=row_number, column=column_index).value


def _validate_quantity(value: Any) -> Tuple[str, str]:
    if value is None:
        return "fail", "Quantity is required and is blank or missing."
    text = _clean_text(value)
    if not text:
        return "fail", "Quantity is required and is blank or missing."
    try:
        quantity = Decimal(text)
    except InvalidOperation:
        return "fail", f"Quantity is not numeric: {text}."
    if quantity < 0:
        return "fail", f"Quantity must not be negative: {text}."
    return "pass", "Quantity passed import validation."


def _normalize_quantity(value: Any) -> Optional[float]:
    try:
        return float(Decimal(_clean_text(value)))
    except (InvalidOperation, ValueError):
        return None


def _readiness_status(validation_status: str) -> str:
    if validation_status == "pass":
        return "import_valid_pricing_blocked"
    if validation_status == "excluded":
        return "excluded"
    if validation_status == "warning":
        return "review_required_pricing_blocked"
    return "blocked"


def _is_excluded(value: Any) -> bool:
    return _normalize_header(_clean_text(value)) in EXCLUSION_MARKERS


def _clean_text(value: Any) -> str:
    if value is None:
        return ""
    return str(value).strip()


def _normalize_header(value: str) -> str:
    return " ".join(_clean_text(value).lower().replace("_", " ").split())


def build_client_boq_intake_review_summary(intake_result: Dict[str, Any]) -> Dict[str, Any]:
    """Summarise an existing intake result for controlled estimator review."""
    rows = list(intake_result.get("rows") or [])
    grouped_rows = {
        "passed": [],
        "blocked": [],
        "warning": [],
        "excluded": [],
        "review_required": [],
    }
    validation_messages = {
        "workbook": list(intake_result.get("workbook_messages") or []),
        "column_mapping": list(intake_result.get("column_mapping_messages") or []),
        "rows": [],
    }

    has_fail = _status_is_fail_closed(intake_result.get("workbook_status")) or _status_is_fail_closed(
        intake_result.get("column_mapping_status")
    )
    has_review = _status_requires_review(intake_result.get("workbook_status")) or _status_requires_review(
        intake_result.get("column_mapping_status")
    )

    for row in rows:
        row_status = _normalize_status(row.get("validation_status"))
        row_ref = _row_reference(row, intake_result)
        row_messages = list(row.get("validation_messages") or [])

        if row_status == "pass":
            grouped_rows["passed"].append(row_ref)
        elif row_status == "warning":
            grouped_rows["warning"].append(row_ref)
            has_review = True
        elif row_status == "excluded":
            grouped_rows["excluded"].append(row_ref)
            has_review = True
        else:
            grouped_rows["blocked"].append(row_ref)
            has_fail = True

        if row.get("estimator_review_required"):
            grouped_rows["review_required"].append(row_ref)
            has_review = True

        if row_messages:
            validation_messages["rows"].append(
                {
                    "row_reference": row_ref,
                    "validation_status": row_status or "blank",
                    "messages": row_messages,
                }
            )

    row_counts = {
        "total_rows": len(rows),
        "passed_rows": len(grouped_rows["passed"]),
        "blocked_rows": len(grouped_rows["blocked"]),
        "warning_rows": len(grouped_rows["warning"]),
        "excluded_rows": len(grouped_rows["excluded"]),
        "review_required_rows": len(grouped_rows["review_required"]),
    }

    summary_status = "fail" if has_fail else "review_required" if has_review else "pass"

    return {
        "source_file_ref": intake_result.get("source_file_ref"),
        "workbook_status": _normalize_status(intake_result.get("workbook_status")),
        "selected_worksheet": intake_result.get("selected_worksheet"),
        "column_mapping_status": _normalize_status(intake_result.get("column_mapping_status")),
        "summary_status": summary_status,
        "row_counts": row_counts,
        "row_references": grouped_rows,
        "validation_messages_for_estimator_review": validation_messages,
        "pricing_approval": False,
        "pricing_ready": False,
        "export_ready": False,
    }


def _row_reference(row: Dict[str, Any], intake_result: Dict[str, Any]) -> Dict[str, Any]:
    return {
        "source_file_ref": row.get("source_file_ref") or intake_result.get("source_file_ref"),
        "source_sheet_name": row.get("source_sheet_name"),
        "source_row_number": row.get("source_row_number"),
    }


def _normalize_status(value: Any) -> str:
    return _normalize_header(_clean_text(value))


def _status_is_fail_closed(value: Any) -> bool:
    status = _normalize_status(value)
    return status in {"", "fail", "failed", "unknown", "not checked", "not_checked"}


def _status_requires_review(value: Any) -> bool:
    status = _normalize_status(value)
    return status in {"warning", "review required", "review_required", "excluded"}


def build_client_boq_intake_review_decision(summary: Dict[str, Any]) -> Dict[str, Any]:
    """Classify an intake review summary for estimator decision routing."""
    reason_codes: List[str] = []
    original_summary_status = summary.get("summary_status")
    summary_status = _normalize_status(original_summary_status)
    row_counts = summary.get("row_counts")
    row_references = summary.get("row_references")
    validation_messages = summary.get("validation_messages_for_estimator_review")

    counts, counts_valid = _decision_row_counts(row_counts)
    row_references_valid = isinstance(row_references, dict)
    validation_messages_valid = isinstance(validation_messages, dict)

    if "summary_status" not in summary or summary_status == "":
        reason_codes.append("summary_status_missing")
    elif summary_status in {"unknown", "not checked", "not_checked"}:
        reason_codes.append("summary_status_unknown")
    elif summary_status in {"fail", "failed"}:
        reason_codes.append("summary_failed")

    if not counts_valid:
        reason_codes.append("row_counts_missing")
    else:
        if counts["blocked_rows"] > 0:
            reason_codes.append("blocked_rows_present")
        if counts["warning_rows"] > 0:
            reason_codes.append("warning_rows_present")
        if counts["excluded_rows"] > 0:
            reason_codes.append("excluded_rows_present")
        if counts["review_required_rows"] > 0:
            reason_codes.append("review_required_rows_present")

    if _status_requires_review(summary.get("column_mapping_status")):
        reason_codes.append("column_mapping_review_required")

    required_evidence_missing = not counts_valid or not row_references_valid or not validation_messages_valid
    summary_failed = summary_status in {"", "fail", "failed", "unknown", "not checked", "not_checked"}
    blocked_rows_present = counts_valid and counts["blocked_rows"] > 0
    excluded_rows_present = counts_valid and counts["excluded_rows"] > 0
    pass_counts_clean = counts_valid and _decision_row_counts_are_clean(counts)

    if summary_failed or required_evidence_missing or blocked_rows_present:
        decision_status = "import_failed_pricing_blocked"
    elif excluded_rows_present:
        decision_status = "excluded_or_partial_review_required"
    elif summary_status == "review required":
        decision_status = "review_required_pricing_blocked"
    elif summary_status == "pass" and pass_counts_clean:
        decision_status = "import_clean_pricing_blocked"
        reason_codes.append("import_clean")
    else:
        decision_status = "import_failed_pricing_blocked"

    _append_decision_blockers(reason_codes)

    return {
        "decision_status": decision_status,
        "reason_codes": reason_codes,
        "summary_status": original_summary_status,
        "row_counts": row_counts,
        "row_references": row_references,
        "validation_messages_for_estimator_review": validation_messages,
        "pricing_approval": False,
        "pricing_ready": False,
        "export_ready": False,
        "client_return_ready": False,
    }


def _decision_row_counts(value: Any) -> Tuple[Dict[str, int], bool]:
    required_keys = (
        "total_rows",
        "passed_rows",
        "blocked_rows",
        "warning_rows",
        "excluded_rows",
        "review_required_rows",
    )
    if not isinstance(value, dict):
        return {}, False

    counts: Dict[str, int] = {}
    for key in required_keys:
        raw_value = value.get(key)
        if not isinstance(raw_value, int) or raw_value < 0:
            return {}, False
        counts[key] = raw_value

    return counts, True


def _decision_row_counts_are_clean(counts: Dict[str, int]) -> bool:
    return (
        counts["blocked_rows"] == 0
        and counts["warning_rows"] == 0
        and counts["excluded_rows"] == 0
        and counts["review_required_rows"] == 0
        and counts["total_rows"] == counts["passed_rows"]
    )


def _append_decision_blockers(reason_codes: List[str]) -> None:
    for code in ("pricing_blocked", "export_blocked", "client_return_blocked"):
        if code not in reason_codes:
            reason_codes.append(code)


def build_client_boq_intake_review_packet(
    summary: Dict[str, Any],
    decision: Dict[str, Any],
) -> Dict[str, Any]:
    """Assemble deterministic estimator-facing review packet evidence."""
    invalid_reason_codes: List[str] = []
    summary_valid = isinstance(summary, dict)
    decision_valid = isinstance(decision, dict)
    summary_data = summary if summary_valid else {}
    decision_data = decision if decision_valid else {}

    if not summary_valid:
        invalid_reason_codes.append("summary_missing")
    if not decision_valid:
        invalid_reason_codes.append("decision_missing")

    summary_status = summary_data.get("summary_status")
    decision_summary_status = decision_data.get("summary_status")
    decision_status = decision_data.get("decision_status")

    if _packet_status_is_missing_or_unknown(summary_status):
        invalid_reason_codes.append("summary_status_missing")
    if _packet_status_is_missing_or_unknown(decision_summary_status):
        invalid_reason_codes.append("decision_summary_status_missing")
    decision_status_key = _packet_decision_status_key(decision_status)
    if decision_status_key == "":
        invalid_reason_codes.append("decision_status_missing")
    elif decision_status_key in {"unknown", "not checked", "not_checked"}:
        invalid_reason_codes.append("decision_status_unknown")
    elif decision_status_key not in _PACKET_STATUS_BY_DECISION_STATUS:
        invalid_reason_codes.append("decision_status_unknown")

    if summary_valid and decision_valid:
        if summary_data.get("summary_status") != decision_data.get("summary_status"):
            invalid_reason_codes.append("summary_decision_status_mismatch")
        if summary_data.get("row_counts") != decision_data.get("row_counts"):
            invalid_reason_codes.append("row_counts_mismatch")
        if summary_data.get("row_references") != decision_data.get("row_references"):
            invalid_reason_codes.append("row_references_mismatch")
        if summary_data.get("validation_messages_for_estimator_review") != decision_data.get(
            "validation_messages_for_estimator_review"
        ):
            invalid_reason_codes.append("validation_messages_mismatch")

    if not _packet_readiness_guards_are_valid(summary_data, decision_data):
        invalid_reason_codes.append("readiness_guard_invalid")

    if invalid_reason_codes:
        review_packet_status = "review_packet_invalid_pricing_blocked"
        reason_codes = invalid_reason_codes + ["packet_invalid"]
    else:
        review_packet_status = _PACKET_STATUS_BY_DECISION_STATUS[decision_status_key]
        reason_codes = list(decision_data.get("reason_codes") or [])

    _append_decision_blockers(reason_codes)

    source_file_ref = summary_data.get("source_file_ref")
    selected_worksheet = summary_data.get("selected_worksheet")
    workbook_status = summary_data.get("workbook_status")
    column_mapping_status = summary_data.get("column_mapping_status")
    row_counts = summary_data.get("row_counts")
    row_references = summary_data.get("row_references")
    validation_messages = summary_data.get("validation_messages_for_estimator_review")

    return {
        "review_packet_status": review_packet_status,
        "source_file_ref": source_file_ref,
        "selected_worksheet": selected_worksheet,
        "workbook_status": workbook_status,
        "column_mapping_status": column_mapping_status,
        "summary_status": summary_status,
        "decision_status": decision_status,
        "reason_codes": reason_codes,
        "row_counts": row_counts,
        "row_references": row_references,
        "validation_messages_for_estimator_review": validation_messages,
        "review_packet_sections": _build_review_packet_sections(
            source_file_ref,
            selected_worksheet,
            workbook_status,
            column_mapping_status,
            summary_status,
            decision_status,
            row_counts,
            row_references,
            validation_messages,
        ),
        "pricing_approval": False,
        "pricing_ready": False,
        "export_ready": False,
        "client_return_ready": False,
    }


_PACKET_STATUS_BY_DECISION_STATUS = {
    "import_clean_pricing_blocked": "review_packet_clean_pricing_blocked",
    "review_required_pricing_blocked": "review_packet_review_required_pricing_blocked",
    "import_failed_pricing_blocked": "review_packet_failed_pricing_blocked",
    "excluded_or_partial_review_required": "review_packet_excluded_or_partial_review_required",
}


def _packet_status_is_missing_or_unknown(value: Any) -> bool:
    status = _normalize_status(value)
    return status in {"", "unknown", "not checked", "not_checked"}


def _packet_decision_status_key(value: Any) -> str:
    return _clean_text(value).lower()


def _packet_readiness_guards_are_valid(
    summary: Dict[str, Any],
    decision: Dict[str, Any],
) -> bool:
    summary_guards = ("pricing_approval", "pricing_ready", "export_ready")
    decision_guards = ("pricing_approval", "pricing_ready", "export_ready", "client_return_ready")
    for key in summary_guards:
        if summary.get(key) is not False:
            return False
    for key in decision_guards:
        if decision.get(key) is not False:
            return False
    return True


def _build_review_packet_sections(
    source_file_ref: Any,
    selected_worksheet: Any,
    workbook_status: Any,
    column_mapping_status: Any,
    summary_status: Any,
    decision_status: Any,
    row_counts: Any,
    row_references: Any,
    validation_messages: Any,
) -> Dict[str, Any]:
    return {
        "source": {
            "source_file_ref": source_file_ref,
            "selected_worksheet": selected_worksheet,
        },
        "statuses": {
            "workbook_status": workbook_status,
            "column_mapping_status": column_mapping_status,
            "summary_status": summary_status,
            "decision_status": decision_status,
        },
        "row_evidence": {
            "row_counts": row_counts,
            "row_references": row_references,
        },
        "messages": {
            "validation_messages_for_estimator_review": validation_messages,
        },
        "readiness_guards": {
            "pricing_approval": False,
            "pricing_ready": False,
            "export_ready": False,
            "client_return_ready": False,
        },
    }
