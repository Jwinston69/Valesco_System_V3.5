import ast
import inspect
import unittest
from pathlib import Path

from openpyxl import load_workbook

import engine.modules.client_boq_intake_v0_1 as intake


FIXTURE_DIR = Path(__file__).parent / "fixtures" / "phase_2" / "client_boq_intake"


def fixture_path(name):
    return FIXTURE_DIR / name


class TestClientBOQIntakeValidationFoundationV01(unittest.TestCase):
    def test_workbook_readability_and_worksheet_identification(self):
        result = intake.import_client_boq_workbook(
            str(fixture_path("valid_minimal_client_boq.xlsx"))
        )

        self.assertTrue(result["workbook_readable"])
        self.assertEqual(result["workbook_status"], "pass")
        self.assertEqual(
            result["sheets"],
            [{"worksheet_name": "Client BOQ", "worksheet_index": 1}],
        )
        self.assertEqual(
            result["selected_worksheet"],
            {"worksheet_name": "Client BOQ", "worksheet_index": 1},
        )

    def test_deterministic_column_mapping(self):
        result = intake.import_client_boq_workbook(
            str(fixture_path("valid_minimal_client_boq.xlsx"))
        )

        self.assertEqual(result["column_mapping_status"], "pass")
        mapping = result["column_mapping"]
        self.assertEqual(mapping["client_item_ref"]["column_letter"], "A")
        self.assertEqual(mapping["description"]["column_letter"], "B")
        self.assertEqual(mapping["unit"]["column_letter"], "C")
        self.assertEqual(mapping["quantity"]["column_letter"], "D")

    def test_source_traceability_and_original_values_are_preserved(self):
        result = intake.import_client_boq_workbook(
            str(fixture_path("valid_minimal_client_boq.xlsx"))
        )
        row = result["rows"][0]

        self.assertEqual(row["source_file_ref"], "valid_minimal_client_boq.xlsx")
        self.assertEqual(row["source_sheet_name"], "Client BOQ")
        self.assertEqual(row["source_row_number"], 2)
        self.assertEqual(row["client_item_ref"], "A1")
        self.assertEqual(row["description_original"], "Artificial wall lining")
        self.assertEqual(row["unit_original"], "m2")
        self.assertEqual(row["quantity_original"], 12)
        self.assertEqual(row["quantity_normalized"], 12.0)

    def test_valid_import_does_not_imply_pricing_or_export_readiness(self):
        result = intake.import_client_boq_workbook(
            str(fixture_path("valid_minimal_client_boq.xlsx"))
        )

        self.assertFalse(result["pricing_approval"])
        self.assertFalse(result["export_ready"])
        for row in result["rows"]:
            self.assertEqual(row["validation_status"], "pass")
            self.assertEqual(row["readiness_status"], "import_valid_pricing_blocked")
            self.assertFalse(row["pricing_ready"])
            self.assertFalse(row["export_ready"])

    def test_blank_required_fields_fail_closed(self):
        result = intake.import_client_boq_workbook(
            str(fixture_path("missing_required_fields_client_boq.xlsx"))
        )

        self.assertEqual(result["workbook_status"], "fail")
        failed_rows = [row for row in result["rows"] if row["validation_status"] == "fail"]
        self.assertEqual(len(failed_rows), 3)
        messages = " ".join(
            message
            for row in failed_rows
            for message in row["validation_messages"]
        )
        self.assertIn("Description is required", messages)
        self.assertIn("Unit is required", messages)
        self.assertIn("Quantity is required", messages)
        for row in failed_rows:
            self.assertEqual(row["readiness_status"], "blocked")
            self.assertTrue(row["estimator_review_required"])
            self.assertFalse(row["pricing_ready"])

    def test_invalid_quantity_fails_closed(self):
        result = intake.import_client_boq_workbook(
            str(fixture_path("invalid_quantity_client_boq.xlsx"))
        )

        self.assertEqual(result["workbook_status"], "fail")
        self.assertEqual([row["validation_status"] for row in result["rows"]], ["fail", "fail"])
        messages = " ".join(
            message
            for row in result["rows"]
            for message in row["validation_messages"]
        )
        self.assertIn("Quantity is not numeric: many.", messages)
        self.assertIn("Quantity must not be negative: -1.", messages)

    def test_ambiguous_mapping_requires_visible_review(self):
        result = intake.import_client_boq_workbook(
            str(fixture_path("ambiguous_mapping_client_boq.xlsx"))
        )

        self.assertEqual(result["column_mapping_status"], "review_required")
        self.assertEqual(result["workbook_status"], "warning")
        self.assertIn(
            "Ambiguous column mapping for description",
            " ".join(result["column_mapping_messages"]),
        )
        self.assertIn(
            "Column mapping requires review before pricing readiness.",
            result["workbook_messages"],
        )
        self.assertFalse(result["pricing_approval"])
        self.assertFalse(result["export_ready"])

    def test_hidden_rows_hidden_columns_and_merged_cells_are_visible(self):
        result = intake.import_client_boq_workbook(
            str(fixture_path("hidden_or_merged_structure_client_boq.xlsx"))
        )

        workbook_messages = " ".join(result["workbook_messages"])
        self.assertIn("Hidden rows visible for review", workbook_messages)
        self.assertIn("Hidden columns visible for review", workbook_messages)
        self.assertIn("Merged cells visible for review", workbook_messages)

        row_messages = {
            row["source_row_number"]: " ".join(row["validation_messages"])
            for row in result["rows"]
        }
        self.assertIn("Source row 2 is hidden", row_messages[2])
        self.assertIn("Mapped quantity column D is hidden", row_messages[3])
        self.assertIn("Mapped description cell B4 is merged", row_messages[4])

        review_rows = [row for row in result["rows"] if row["source_row_number"] in {2, 3}]
        for row in review_rows:
            self.assertEqual(row["validation_status"], "warning")
            self.assertEqual(row["readiness_status"], "review_required_pricing_blocked")
            self.assertTrue(row["estimator_review_required"])
            self.assertFalse(row["pricing_ready"])

        merged_row = [row for row in result["rows"] if row["source_row_number"] == 4][0]
        self.assertEqual(merged_row["validation_status"], "fail")
        self.assertEqual(merged_row["readiness_status"], "blocked")
        self.assertTrue(merged_row["estimator_review_required"])
        self.assertFalse(merged_row["pricing_ready"])

    def test_excluded_rows_have_visible_reasons(self):
        result = intake.import_client_boq_workbook(
            str(fixture_path("hidden_or_merged_structure_client_boq.xlsx"))
        )

        excluded = [row for row in result["rows"] if row["validation_status"] == "excluded"]
        self.assertEqual(len(excluded), 1)
        self.assertEqual(excluded[0]["source_row_number"], 5)
        self.assertEqual(excluded[0]["exclusion_reason"], "Fixture row marked as excluded.")
        self.assertIn("Fixture row marked as excluded.", excluded[0]["validation_messages"])
        self.assertEqual(excluded[0]["readiness_status"], "excluded")
        self.assertFalse(excluded[0]["pricing_ready"])

    def test_artificial_fixtures_contain_no_real_client_tender_or_rate_data(self):
        fixture_names = {
            "valid_minimal_client_boq.xlsx",
            "missing_required_fields_client_boq.xlsx",
            "invalid_quantity_client_boq.xlsx",
            "ambiguous_mapping_client_boq.xlsx",
            "hidden_or_merged_structure_client_boq.xlsx",
        }
        self.assertEqual({path.name for path in FIXTURE_DIR.glob("*.xlsx")}, fixture_names)

        prohibited_terms = {
            "client",
            "tender",
            "supplier",
            "quotation",
            "quote",
            "project",
            "rate",
            "return",
            "james",
            "valesco",
        }

        for workbook_path in sorted(FIXTURE_DIR.glob("*.xlsx")):
            workbook = load_workbook(workbook_path, data_only=False, read_only=False)
            for worksheet in workbook.worksheets:
                for row in worksheet.iter_rows():
                    for cell in row:
                        if cell.value is None:
                            continue
                        value = str(cell.value).lower()
                        for term in prohibited_terms:
                            self.assertNotIn(
                                term,
                                value,
                                f"Prohibited term {term!r} found in {workbook_path.name}",
                            )

    def test_module_has_no_prohibited_pipeline_imports(self):
        source = inspect.getsource(intake)
        tree = ast.parse(source)
        imports = set()

        for node in ast.walk(tree):
            if isinstance(node, ast.Import):
                for alias in node.names:
                    imports.add(alias.name)
            elif isinstance(node, ast.ImportFrom) and node.module:
                imports.add(node.module)
                for alias in node.names:
                    imports.add(f"{node.module}.{alias.name}")

        forbidden = {
            "engine.modules.architect_v2_1",
            "engine.modules.ce_backend_adapter_v3_2",
            "engine.modules.ce_retrieval_layer_v2_1",
            "engine.modules.estimator_runtime_v2_1",
            "engine.modules.merge_agent_v2_1",
            "engine.modules.pricing_engine_v3_4",
            "engine.modules.pricing_logic_v2_1",
            "engine.modules.router_v2_1",
            "engine.modules.validator_v2_1",
            "engine.scripts.mvp_runner_v2_1",
            "engine.scripts.mvp_runner_v2_2",
        }

        for module_name in sorted(imports):
            for blocked in forbidden:
                self.assertFalse(
                    module_name == blocked or module_name.startswith(f"{blocked}."),
                    f"Forbidden dependency detected: {module_name}",
                )


if __name__ == "__main__":
    unittest.main()
